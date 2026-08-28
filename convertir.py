import json
import os
from datetime import datetime, date
from openpyxl import load_workbook


MESES = [
    "Ene", "Feb", "Mar", "Abr", "May", "Jun",
    "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"
]


# =========================================================
# CONVERTIR VALORES NUMÉRICOS
# =========================================================

def numero(valor):
    if valor is None or valor == "":
        return 0

    if isinstance(valor, (int, float)):
        return valor

    texto = str(valor).strip().replace("%", "").replace(" ", "")

    if not texto:
        return 0

    try:
        if "," in texto and "." in texto:
            if texto.rfind(",") > texto.rfind("."):
                texto = texto.replace(".", "").replace(",", ".")
            else:
                texto = texto.replace(",", "")

        elif "," in texto:
            texto = texto.replace(",", ".")

        return float(texto)

    except Exception:
        return 0


def entero(valor):
    return round(numero(valor))


# =========================================================
# CONVERTIR FECHAS
# =========================================================

def fecha(valor):

    if valor is None or valor == "":
        return None

    if isinstance(valor, datetime):
        return date(valor.year, valor.month, valor.day)

    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()

    formatos = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%m/%d/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S"
    ]

    for formato in formatos:
        try:
            return datetime.strptime(texto, formato).date()
        except Exception:
            pass

    return None


# =========================================================
# BUSCAR ENCABEZADOS
# =========================================================

def buscar_encabezados(ws, requeridos):

    for numero_fila, fila in enumerate(
        ws.iter_rows(max_row=15, values_only=True),
        start=1
    ):

        valores = [
            str(valor or "").strip()
            for valor in fila
        ]

        if all(campo in valores for campo in requeridos):

            columnas = {
                valor: indice + 1
                for indice, valor in enumerate(valores)
                if valor
            }

            print(f"Encabezados encontrados en fila {numero_fila}")

            return numero_fila, columnas

    raise Exception(
        "No se encontró la fila de encabezados esperada."
    )


# =========================================================
# LEER FILAS
# =========================================================

def filas(ws, fila_encabezado, columnas):

    resultado = []

    # IMPORTANTE:
    # No usamos ws.max_row porque en read_only puede devolver None.
    for fila_excel, valores_fila in enumerate(
        ws.iter_rows(
            min_row=fila_encabezado + 1,
            values_only=True
        ),
        start=fila_encabezado + 1
    ):

        registro = {}

        for nombre, columna in columnas.items():

            indice = columna - 1

            if indice < len(valores_fila):
                registro[nombre] = valores_fila[indice]
            else:
                registro[nombre] = None

        if any(
            valor is not None and valor != ""
            for valor in registro.values()
        ):
            resultado.append(registro)

    return resultado


# =========================================================
# RESUMEN
# =========================================================

def resumen(registros):

    cantidad = len(registros) or 1

    return {

        "impressions": entero(
            sum(
                numero(r.get("Impresiones (totales)"))
                for r in registros
            )
        ),

        "clicks": entero(
            sum(
                numero(r.get("Clics (totales)"))
                for r in registros
            )
        ),

        "rate": (
            sum(
                numero(
                    r.get("Tasa de interacción (total)")
                )
                for r in registros
            )
            / cantidad
        ),

        "conversions": 0,

        "days": len(registros),

        "reactions": entero(
            sum(
                numero(r.get("Reacciones (total)"))
                for r in registros
            )
        ),

        "comments": entero(
            sum(
                numero(r.get("Comentarios (totales)"))
                for r in registros
            )
        ),

        "shares": entero(
            sum(
                numero(r.get("Veces compartido (total)"))
                for r in registros
            )
        )
    }


# =========================================================
# BUSCAR AUTOMÁTICAMENTE EL EXCEL
# =========================================================

print("====================================")
print("Iniciando conversión del Excel")
print("====================================")


archivos = [
    nombre
    for nombre in os.listdir("reporte")
    if nombre.lower().endswith(
        (".xlsx", ".xlsm")
    )
]


if not archivos:

    raise Exception(
        "No se encontró ningún archivo Excel dentro de reporte/."
    )


# Usar el archivo modificado más recientemente

archivos.sort(
    key=lambda nombre: os.path.getmtime(
        os.path.join("reporte", nombre)
    ),
    reverse=True
)


archivo = os.path.join(
    "reporte",
    archivos[0]
)


print(f"Excel encontrado: {archivo}")


# =========================================================
# ABRIR EXCEL
# =========================================================

wb = load_workbook(
    archivo,
    data_only=True,
    read_only=True
)


print("Hojas encontradas:")
print(wb.sheetnames)


# =========================================================
# HOJA INDICADORES
# =========================================================

if "Indicadores" not in wb.sheetnames:

    raise Exception(
        'No se encontró la hoja "Indicadores".'
    )


ws = wb["Indicadores"]


print("------------------------------------")
print("Procesando hoja: Indicadores")
print("------------------------------------")


fila, columnas = buscar_encabezados(
    ws,
    [
        "Fecha",
        "Impresiones (totales)",
        "Clics (totales)",
        "Tasa de interacción (total)"
    ]
)


print(f"Fila de encabezados: {fila}")
print(f"Columnas encontradas: {columnas}")


indicadores = filas(
    ws,
    fila,
    columnas
)


# =========================================================
# FILTRAR FECHAS VÁLIDAS
# =========================================================

indicadores_validos = []


for registro in indicadores:

    fecha_registro = fecha(
        registro.get("Fecha")
    )

    if fecha_registro is not None:

        registro["Fecha"] = fecha_registro

        indicadores_validos.append(
            registro
        )


indicadores = indicadores_validos


indicadores.sort(
    key=lambda r: r["Fecha"]
)


if not indicadores:

    raise Exception(
        "La hoja Indicadores no contiene fechas válidas."
    )


print(
    f"Registros de indicadores encontrados: "
    f"{len(indicadores)}"
)


# =========================================================
# RESUMEN ANUAL
# =========================================================

years = sorted(
    set(
        registro["Fecha"].year
        for registro in indicadores
    )
)


annual = {}


for year in years:

    registros_year = [
        r
        for r in indicadores
        if r["Fecha"].year == year
    ]

    annual[str(year)] = resumen(
        registros_year
    )


# Resumen de todos los años

annual["all"] = resumen(
    indicadores
)


# =========================================================
# RESUMEN MENSUAL
# =========================================================

grupos = {}


for registro in indicadores:

    year = registro["Fecha"].year
    month = registro["Fecha"].month

    clave = f"{year}-{month:02d}"

    if clave not in grupos:
        grupos[clave] = []

    grupos[clave].append(
        registro
    )


months = []


for clave in sorted(grupos.keys()):

    registros = grupos[clave]

    year, month = map(
        int,
        clave.split("-")
    )

    cantidad = len(registros) or 1


    # Interacciones totales

    interacciones = sum(

        numero(
            r.get("Reacciones (total)")
        )

        +

        numero(
            r.get("Comentarios (totales)")
        )

        +

        numero(
            r.get("Veces compartido (total)")
        )

        for r in registros
    )


    months.append([

        f"{MESES[month - 1]} {str(year)[-2:]}",

        year,

        entero(
            sum(
                numero(
                    r.get("Impresiones (totales)")
                )
                for r in registros
            )
        ),

        entero(
            sum(
                numero(
                    r.get("Clics (totales)")
                )
                for r in registros
            )
        ),

        entero(
            interacciones
        ),

        entero(
            sum(
                numero(
                    r.get("Comentarios (totales)")
                )
                for r in registros
            )
        ),

        entero(
            sum(
                numero(
                    r.get("Veces compartido (total)")
                )
                for r in registros
            )
        ),

        sum(
            numero(
                r.get(
                    "Tasa de interacción (total)"
                )
            )
            for r in registros
        )
        / cantidad

    ])


# =========================================================
# HOJA TODAS LAS PUBLICACIONES
# =========================================================

posts = []


if "Todas las publicaciones" in wb.sheetnames:

    print("------------------------------------")
    print("Procesando hoja: Todas las publicaciones")
    print("------------------------------------")


    ws_posts = wb[
        "Todas las publicaciones"
    ]


    fila_posts, columnas_posts = buscar_encabezados(
        ws_posts,
        [
            "Título de la publicación",
            "Fecha de creación",
            "Impresiones",
            "Clics",
            "Tasa de interacción"
        ]
    )


    publicaciones = filas(
        ws_posts,
        fila_posts,
        columnas_posts
    )


    print(
        f"Publicaciones encontradas: "
        f"{len(publicaciones)}"
    )


    for publicacion in publicaciones:

        titulo = (
            publicacion.get(
                "Título de la publicación"
            )
            or ""
        )


        if not titulo:
            continue


        fecha_publicacion = fecha(
            publicacion.get(
                "Fecha de creación"
            )
        )


        if fecha_publicacion:

            fecha_texto = (
                fecha_publicacion.strftime(
                    "%d/%m/%Y"
                )
            )

        else:

            fecha_texto = str(
                publicacion.get(
                    "Fecha de creación"
                )
                or ""
            )


        posts.append([

            titulo,

            fecha_texto,

            publicacion.get(
                "Tipo de publicación"
            )
            or "",

            entero(
                publicacion.get(
                    "Impresiones"
                )
            ),

            entero(
                publicacion.get(
                    "Clics"
                )
            ),

            numero(
                publicacion.get(
                    "Tasa de interacción"
                )
            )

        ])


else:

    print(
        'Aviso: no existe la hoja '
        '"Todas las publicaciones".'
    )


# =========================================================
# CREAR JSON
# =========================================================

resultado = {

    "annual": annual,

    "months": months,

    "posts": posts

}


with open(
    "data.json",
    "w",
    encoding="utf-8"
) as archivo_json:

    json.dump(
        resultado,
        archivo_json,
        ensure_ascii=False,
        indent=2
    )


# =========================================================
# CERRAR EXCEL
# =========================================================

wb.close()


# =========================================================
# RESUMEN FINAL
# =========================================================

print("")
print("====================================")
print("CONVERSIÓN COMPLETADA CORRECTAMENTE")
print("====================================")

print(f"Archivo: {archivo}")

print(f"Años: {years}")

print(
    f"Días/registros: "
    f"{len(indicadores)}"
)

print(
    f"Meses: "
    f"{len(months)}"
)

print(
    f"Publicaciones: "
    f"{len(posts)}"
)

print(
    f"Impresiones: "
    f"{annual['all']['impressions']}"
)

print(
    f"Clics: "
    f"{annual['all']['clicks']}"
)

print(
    f"Interacción: "
    f"{annual['all']['rate']}"
)

print("====================================")
